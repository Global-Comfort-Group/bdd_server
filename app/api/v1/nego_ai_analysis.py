import io
import csv
import json
from datetime import datetime, date as date_type
from typing import Optional, List

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from pydantic import BaseModel, field_validator
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.auth import get_current_user
from app.core.config import settings
from app.core.database import get_async_session
from app.models.negotiation_chronicle import NegotiationChronicleAttachment
from app.models.nego_table import NegoTable
from app.models.property import Property
from app.models.user import User
from app.schemas.negotiation_chronicle import NegotiationChronicleAttachment as NegotiationChronicleAttachmentSchema
from app.services.file_storage import file_storage_service
from sqlalchemy import select

router = APIRouter()


# --------------------------------------------------------------------------- #
# Response schema                                                              #
# --------------------------------------------------------------------------- #

class PriceComparison(BaseModel):
    askingPrice: Optional[str] = None
    clientOffer: Optional[str] = None
    gap: Optional[str] = None
    currency: Optional[str] = None


class RecentTransaction(BaseModel):
    date: Optional[str] = None
    title: str                    # short 3-6 word label e.g. "Cash Payment Agreed"
    summary: str
    event_type: Optional[str] = None  # payment_method|price_change|offer|agreement|condition|meeting
    outcome: Optional[str] = None


class RawExtracted(BaseModel):
    header: str
    value: str


class AIAnalysisResult(BaseModel):
    useful: bool
    reason: Optional[str] = None
    negotiationStatus: Optional[str] = None
    priceComparison: Optional[PriceComparison] = None
    clientConditions: Optional[List[str]] = None
    recentTransactions: Optional[List[RecentTransaction]] = None
    keyNotes: Optional[List[str]] = None
    rawExtracted: Optional[List[RawExtracted]] = None


class UploadAnalysisResponse(BaseModel):
    attachment_id: int
    ai_result: AIAnalysisResult


# New comparison schema

class NegotiationItem(BaseModel):
    title: str
    bdd_offer: str = ""
    client_offer: str = ""
    status: str = "negotiating"  # "agreed" | "negotiating"
    agreed_date: Optional[str] = None
    final_value: Optional[str] = None
    notes: Optional[str] = None

    @field_validator('title', 'bdd_offer', 'client_offer', mode='before')
    @classmethod
    def coerce_none_to_str(cls, v: object) -> str:
        return "" if v is None else str(v)


class AIComparisonResult(BaseModel):
    useful: bool
    reason: Optional[str] = None
    overall_status: Optional[str] = None
    summary: Optional[str] = None
    negotiation_items: List[NegotiationItem] = []


class UploadForPropertyResponse(BaseModel):
    attachment_id: int
    ai_result: AIComparisonResult
    nego_table_id: int


# --------------------------------------------------------------------------- #
# Post-processing: enforce matching-value = agreed                            #
# --------------------------------------------------------------------------- #

def _normalize_for_compare(value: str) -> str:
    """
    Normalize a negotiation value for equality comparison.
    Handles: currency symbols, thousands separators, M/K shorthand, unit
    variations, and trailing formatting so that, e.g.:
      '₱625,000' == '625000' == 'PHP 625,000'
      '5M'       == '5,000,000'
      '/month'   == 'per month' == 'monthly'
      '3 years'  == '3 yrs'
    """
    import re
    if not value:
        return ""
    v = value.strip().lower()
    # Remove currency symbols (₱ $ € £ ¥ and "PHP"/"USD" prefix/suffix)
    v = re.sub(r'[₱$€£¥]', '', v)
    v = re.sub(r'\b(php|usd|eur|sgd)\b', '', v)
    # Expand shorthand multipliers before stripping commas
    #   5m / 5M → 5000000,  1.5m → 1500000,  500k → 500000
    v = re.sub(r'(\d+(?:\.\d+)?)\s*m\b', lambda m: str(int(float(m.group(1)) * 1_000_000)), v)
    v = re.sub(r'(\d+(?:\.\d+)?)\s*k\b', lambda m: str(int(float(m.group(1)) * 1_000)), v)
    # Remove thousands-separator commas (1,500 → 1500)
    v = re.sub(r'(?<=\d),(?=\d{3})', '', v)
    # Normalize trailing .0 on integers (625000.0 → 625000)
    v = re.sub(r'\b(\d+)\.0+\b', r'\1', v)
    # Normalize time/period units to a canonical form
    v = re.sub(r'\bper\s+month\b|\bmonthly\b', '/month', v)
    v = re.sub(r'\bper\s+year\b|\bannually\b|\bper\s+annum\b', '/year', v)
    v = re.sub(r'\byears?\b|\byrs?\b', 'year', v)
    v = re.sub(r'\bmonths?\b|\bmos?\b', 'month', v)
    # Strip time component from datetime strings (2021-07-13 00:00:00 → 2021-07-13)
    v = re.sub(r'^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}(:\d{2})?$', r'\1', v)
    # Collapse whitespace
    v = re.sub(r'\s+', ' ', v)
    # Strip trailing punctuation
    v = v.strip(".,;: ")
    return v


def _deduplicate_items(result: AIComparisonResult) -> AIComparisonResult:
    """
    Merge rows the AI created for the same topic under different labels
    (e.g. 'Security Deposit (Initial Offer)' + 'Security Deposit (Latest Position)')
    by stripping parenthetical suffixes and keeping the last occurrence (latest round).
    """
    import re

    def _base_title(title: str) -> str:
        # Strip trailing parenthetical like "(Initial Offer)", "(Latest Position)", "(Round 1)"
        return re.sub(r'\s*\(.*?\)\s*$', '', title).strip().lower()

    seen: dict = {}  # base_title → index in deduped list
    deduped = []

    for item in result.negotiation_items:
        base = _base_title(item.title)
        if base in seen:
            # Replace the earlier row with this later one (later = more recent round)
            deduped[seen[base]] = item
        else:
            seen[base] = len(deduped)
            deduped.append(item)

    result.negotiation_items = deduped
    return result


def _enforce_agreed_on_match(result: AIComparisonResult) -> AIComparisonResult:
    """
    Two-rule post-processing guard:

    Rule 1 — CONFIRM AGREED (values match):
      If bdd_offer and client_offer normalize to the same string, force status="agreed"
      regardless of what the AI said.  The AI sometimes misses obvious matches.

    Rule 2 — PREVENT HALLUCINATED AGREEMENT (values clearly differ):
      If the AI says "agreed" but values normalize to DIFFERENT strings, override to
      "negotiating" and wipe final_value / agreed_date.
      We do NOT override when the AI already says "negotiating" and values differ —
      that is a consistent, correct state and touching it would create contradictions
      between the AI's own summary and the item statuses.
    """
    for item in result.negotiation_items:
        bdd = _normalize_for_compare(item.bdd_offer)
        client = _normalize_for_compare(item.client_offer)

        if not bdd or not client:
            # One side is missing — cannot determine, leave AI decision as-is
            continue

        if bdd == client:
            # Rule 1: values match → always agreed
            item.status = "agreed"
            if not item.final_value:
                item.final_value = item.bdd_offer
        elif item.status == "agreed":
            # Rule 2: AI claims agreed but normalized values differ → hallucination, fix it
            item.status = "negotiating"
            item.final_value = None
            item.agreed_date = None
        # else: AI says "negotiating" and values differ → consistent, leave untouched

    # Strip time component from agreed_date (e.g. "2021-07-13 00:00:00" → "2021-07-13")
    import re as _re2
    for item in result.negotiation_items:
        if item.agreed_date:
            item.agreed_date = _re2.sub(r'\s+\d{2}:\d{2}(:\d{2})?$', '', item.agreed_date).strip()

    # For items still being negotiated, fill the date field with "Negotiating"
    # so the frontend table always has something in that column (never blank).
    for item in result.negotiation_items:
        if item.status == "negotiating" and not item.agreed_date:
            item.agreed_date = "Negotiating"

    # Recompute overall_status based on enforced statuses
    if result.negotiation_items:
        total = len(result.negotiation_items)
        agreed = sum(1 for i in result.negotiation_items if i.status == "agreed")
        if agreed == total:
            result.overall_status = "Deal Closed – All terms agreed"
        elif agreed == 0:
            result.overall_status = "All terms still under negotiation"
        else:
            still_open = [i.title for i in result.negotiation_items if i.status == "negotiating"]
            result.overall_status = (
                f"{agreed} of {total} items agreed – "
                f"{', '.join(still_open)} still negotiating"
            )

    return result


def _strip_hallucinated_dates(result: AIComparisonResult, sheet_text: str) -> AIComparisonResult:
    """
    Post-processing guard: null out any agreed_date whose year does not appear
    anywhere in the raw sheet text.  If the spreadsheet only contains 2025/2026,
    a model-fabricated "2023-07-13" is silently dropped rather than shown to the user.
    """
    import re as _re3

    # Collect every 4-digit year that actually appears in the spreadsheet
    sheet_years: set[str] = set(_re3.findall(r'\b(20\d{2})\b', sheet_text))

    for item in result.negotiation_items:
        if not item.agreed_date:
            continue
        date_years = set(_re3.findall(r'\b(20\d{2})\b', item.agreed_date))
        if date_years and not date_years.intersection(sheet_years):
            # The AI invented a year that is not in the spreadsheet — wipe it
            item.agreed_date = None

    return result


def _strip_hallucinated_transaction_dates(result: "AIAnalysisResult", sheet_text: str) -> "AIAnalysisResult":
    """
    Same guard for the /analyze endpoint's recentTransactions dates.
    """
    import re as _re4

    sheet_years: set[str] = set(_re4.findall(r'\b(20\d{2})\b', sheet_text))
    if not result.recentTransactions:
        return result

    for tx in result.recentTransactions:
        if not tx.date:
            continue
        date_years = set(_re4.findall(r'\b(20\d{2})\b', tx.date))
        if date_years and not date_years.intersection(sheet_years):
            tx.date = None

    return result


# --------------------------------------------------------------------------- #
# Sheet structure detector                                                     #
# --------------------------------------------------------------------------- #

def _detect_sheet_structure(content: bytes, filename: str, target_sheet_hint: str = "") -> str:
    """
    Scan the first ~20 rows of the spreadsheet and build a plain-English
    'DETECTED STRUCTURE' block that describes:
      - Which columns appear to be Owner/Client-side vs BDD/Company-side
      - Which rows appear to be date/round headers
      - How many negotiation rounds were found
    This is injected at the top of the prompt so the AI has a structural
    guide before it reads the raw cell dump.
    """
    import re as _re

    lower = filename.lower()
    if not lower.endswith(".xlsx") and not lower.endswith(".xls"):
        return ""  # CSV: flat structure, no need for a guide

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return ""

    # If a hint is provided, only analyse the matching sheet
    target = _find_best_sheet(list(wb.sheetnames), target_sheet_hint)
    sheets_to_check = [target] if target else list(wb.sheetnames)[:3]

    lines: list[str] = []
    for sheet_name in sheets_to_check:
        ws = wb[sheet_name]
        grid: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(max_row=30, values_only=True)):
            cells = []
            for c in row[:50]:
                if c is None:
                    cells.append("")
                elif isinstance(c, datetime):
                    cells.append(c.strftime("%Y-%m-%d"))
                elif isinstance(c, date_type):
                    cells.append(c.strftime("%Y-%m-%d"))
                else:
                    cells.append(str(c).strip())
            grid.append(cells)
            if i >= 29:
                break

        if not grid:
            continue

        # Detect Owner/BDD side-header row
        owner_cols: list[int] = []
        bdd_cols: list[int] = []
        side_row_idx = -1
        date_row_idx = -1

        for ri, row in enumerate(grid):
            owner_hits = [(ci, c) for ci, c in enumerate(row) if _re.search(r'\b(owner|client|lessee|buyer)\b', c.lower())]
            bdd_hits = [(ci, c) for ci, c in enumerate(row) if _re.search(r'\b(company|bdd|broker|lessor|seller)\b', c.lower())]
            if len(owner_hits) >= 2 and len(bdd_hits) >= 2:
                side_row_idx = ri
                owner_cols = [ci for ci, _ in owner_hits]
                bdd_cols = [ci for ci, _ in bdd_hits]
                break

        # Detect date row: look within 5 rows above side_row for a row that has
        # ≥2 REAL date values — a real date must contain a 20XX year OR a month name
        # paired with a year.  Plain 4-digit sequences like phone numbers are excluded.
        def _is_real_date(v: str) -> bool:
            if not v:
                return False
            # ISO date or year 20XX
            if _re.search(r'\b20\d{2}\b', v):
                return True
            # Month name (written out) — only count if it also has a digit nearby
            if _re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\b', v.lower()) \
                    and _re.search(r'\d', v):
                return True
            # MM/DD/YYYY or DD/MM/YYYY style
            if _re.search(r'\b\d{1,2}/\d{1,2}/\d{4}\b', v):
                return True
            return False

        if side_row_idx > 0:
            for ri in range(max(0, side_row_idx - 5), side_row_idx):
                date_count = sum(1 for c in grid[ri] if _is_real_date(c))
                if date_count >= 2:
                    date_row_idx = ri
                    break

        if side_row_idx < 0:
            continue  # no negotiation structure found

        # Pair owner and BDD columns — they should alternate in the same row.
        # Cap at 15 rounds to avoid bloating the prompt with noise.
        num_rounds = min(len(owner_cols), len(bdd_cols), 15)
        lines.append(f"Sheet '{sheet_name}':")
        lines.append(f"  - Negotiation header row: row {side_row_idx + 1}")
        lines.append(f"  - Rounds detected: {num_rounds}")
        lines.append(f"  - Owner/Client columns (1-indexed): {[c + 1 for c in owner_cols[:num_rounds]]}")
        lines.append(f"  - BDD/Company columns (1-indexed): {[c + 1 for c in bdd_cols[:num_rounds]]}")

        if date_row_idx >= 0:
            date_labels = [grid[date_row_idx][oc] for oc in owner_cols[:num_rounds] if oc < len(grid[date_row_idx])]
            # Only include values that are genuine dates
            date_labels = [d for d in date_labels if _is_real_date(d)]
            if date_labels:
                lines.append(f"  - Round dates: {date_labels}")
            else:
                lines.append(f"  - Round dates: not found in header rows (look for dates in the raw data)")
        lines.append(f"  - Columns go LEFT→RIGHT = oldest round → most recent round")

    wb.close()

    if not lines:
        return ""
    return "=== DETECTED SPREADSHEET STRUCTURE (use this as a reading guide) ===\n" + "\n".join(lines) + "\n\n"


# --------------------------------------------------------------------------- #
# File → plain-text helpers                                                   #
# --------------------------------------------------------------------------- #

def _cell_to_str(cell) -> str:
    if cell is None:
        return ""
    # Handle datetime/date objects from Excel date-formatted cells
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d")
    if isinstance(cell, date_type):
        return cell.strftime("%Y-%m-%d")
    # For numeric cells: detect rate/percentage values (0 < val < 1, ≤ 4 decimal places)
    if isinstance(cell, float) and 0 < cell < 1:
        # Round to avoid floating point noise, display as percentage
        pct = round(cell * 100, 4)
        formatted = f"{int(pct)}%" if pct == int(pct) else f"{pct}%"
        return formatted
    v = str(cell).strip()
    # Convert float-like integers (e.g. "250.0" → "250")
    if v.endswith(".0") and v[:-2].lstrip("-").isdigit():
        v = v[:-2]
    return v


def _extract_xlsx_text(content: bytes, property_name: str = "") -> str:
    """
    Extracts every cell verbatim, with one critical improvement over read_only mode:
    merged cells are expanded so every cell in a merged range carries the value
    (not just the top-left corner cell).

    When property_name is provided (e.g. "Maxx Hotel"), only the sheet whose name
    best matches that property is extracted.  This is essential for multi-property
    workbooks (e.g. 16-sheet files) so the AI only sees the relevant sheet and
    does not read data from a different property.

    This is essential for typical nego Excels where date headers like "Jan 2025"
    are merged across the Owner + BDD columns — without expansion, those dates
    appear only once and the AI cannot tell which round they belong to.

    Output is a pipe-separated table with explicit column numbers so the AI
    can count columns without ambiguity:
      [R1]  Col1: value | Col2: value | Col3: value ...
    """
    # Load in normal mode (not read_only) so merged_cells.ranges is available
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        # Corrupted or password-protected — fall back to read_only
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    # Decide which sheets to process
    target_sheet = _find_best_sheet(wb.sheetnames, property_name)
    if target_sheet:
        sheets_to_process = [target_sheet]
        print(f"📋 Sheet match: property='{property_name}' → sheet='{target_sheet}'")
    else:
        # No hint or no match — use active sheet first, then others
        active_name = wb.active.title if wb.active else None
        if active_name and active_name in wb.sheetnames:
            sheets_to_process = [active_name] + [s for s in wb.sheetnames if s != active_name]
        else:
            sheets_to_process = wb.sheetnames
        print(f"📋 No sheet match for '{property_name}' — processing all {len(sheets_to_process)} sheets")

    sheets_text: list[str] = []

    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]

        # ------------------------------------------------------------------ #
        # Build a 2-D grid with merged-cell values expanded
        # ------------------------------------------------------------------ #
        max_row = min(ws.max_row or 500, 500)
        max_col = min(ws.max_column or 50, 50)

        # Fill grid with plain cell values
        grid: list[list[str]] = [[""] * max_col for _ in range(max_row)]
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                r, c = cell.row - 1, cell.column - 1  # convert to 0-based
                if 0 <= r < max_row and 0 <= c < max_col:
                    grid[r][c] = _cell_to_str(cell.value)

        # Expand merged cell values across the full merge range
        if hasattr(ws, 'merged_cells'):
            for merge_range in ws.merged_cells.ranges:
                min_r = merge_range.min_row - 1
                min_c = merge_range.min_col - 1
                # Skip if the merge origin is outside the capped grid
                if min_r >= max_row or min_c >= max_col:
                    continue
                top_val = grid[min_r][min_c]
                for r in range(min_r, min(merge_range.max_row, max_row)):
                    for c in range(min_c, min(merge_range.max_col, max_col)):
                        grid[r][c] = top_val

        # ------------------------------------------------------------------ #
        # Emit rows as  Col1: value | Col2: value  (skip fully-empty rows)
        # ------------------------------------------------------------------ #
        rows_text: list[str] = []
        for ri, row in enumerate(grid):
            # Trim trailing empty cells
            trimmed = row[:]
            while trimmed and not trimmed[-1]:
                trimmed.pop()
            if not any(trimmed):
                continue
            # Pipe-separated with 1-based column labels
            parts = [f"Col{ci + 1}: {v}" for ci, v in enumerate(trimmed) if v]
            rows_text.append(f"[R{ri + 1}]  " + " | ".join(parts))

        if rows_text:
            sheets_text.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text))

    wb.close()
    return "\n\n".join(sheets_text)


def _UNUSED_extract_xlsx_text_structured(content: bytes) -> str:
    """
    Legacy smart extraction kept for reference — was causing hallucinations
    because the metadata filters stripped real data, leaving the AI with
    sparse input it would fill by inventing content.
    """
    import re as _re

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets_text: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Read all rows into a 2-D list
        grid: list[list[str]] = []
        for i, row in enumerate(ws.iter_rows(max_row=300, values_only=True)):
            cells = [_cell_to_str(c) for c in row[:30]]
            grid.append(cells)
            if i >= 299:
                break

        if not grid:
            continue

        num_cols = max(len(r) for r in grid)
        # Pad all rows to same width
        for r in grid:
            r += [""] * (num_cols - len(r))

        lines: list[str] = []

        # ------------------------------------------------------------------ #
        # 1. Detect the "Owner Side / Company Side" header row
        #    This row alternates those two labels across many columns.
        # ------------------------------------------------------------------ #
        side_row_idx: int = -1
        date_row_idx: int = -1
        owner_cols: list[int] = []   # column indices for Owner/Client side
        company_cols: list[int] = [] # column indices for Company/BDD side

        for ri, row in enumerate(grid):
            owner_hits = sum(1 for c in row if "owner" in c.lower() or "client" in c.lower())
            company_hits = sum(1 for c in row if "company" in c.lower() or "bdd" in c.lower())
            if owner_hits >= 2 and company_hits >= 2:
                side_row_idx = ri
                for ci, c in enumerate(row):
                    cl = c.lower()
                    if "owner" in cl or "client" in cl:
                        owner_cols.append(ci)
                    elif "company" in cl or "bdd" in cl:
                        company_cols.append(ci)
                break

        # Find dates row (row just before side_row, or detect by date pattern)
        if side_row_idx > 0:
            for ri in range(side_row_idx - 1, max(side_row_idx - 4, -1), -1):
                date_count = sum(
                    1 for c in grid[ri]
                    if c and (
                        _re.search(r'\b20\d{2}\b', c) or
                        (_re.search(r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\b', c.lower()) and _re.search(r'\d', c)) or
                        _re.search(r'\b\d{1,2}/\d{1,2}/\d{4}\b', c)
                    )
                )
                if date_count >= 2:
                    date_row_idx = ri
                    break

        # ------------------------------------------------------------------ #
        # 2. Context block (rows above the negotiation table)
        # ------------------------------------------------------------------ #
        context_end = date_row_idx if date_row_idx >= 0 else (side_row_idx if side_row_idx >= 0 else 18)
        context_lines: list[str] = []
        for ri in range(min(context_end, len(grid))):
            first_val = grid[ri][0] if grid[ri] else ""
            if first_val:
                context_lines.append(first_val)
        if context_lines:
            lines.append("=== PROPERTY / DEAL CONTEXT ===")
            lines.extend(context_lines)
            lines.append("")

        # ------------------------------------------------------------------ #
        # 3. If paired structure detected, build structured negotiation table
        # ------------------------------------------------------------------ #
        if side_row_idx >= 0 and owner_cols and company_cols:
            # Build round list: pair owner_col[i] with company_col[i]
            rounds: list[tuple[int, int, str]] = []  # (owner_ci, company_ci, date_label)
            dates_row = grid[date_row_idx] if date_row_idx >= 0 else [""] * num_cols
            for i in range(min(len(owner_cols), len(company_cols))):
                oc, cc = owner_cols[i], company_cols[i]
                # Find date: look in dates_row at or near oc
                date_label = ""
                for dc in range(max(0, oc - 1), min(num_cols, oc + 2)):
                    if dates_row[dc]:
                        date_label = dates_row[dc]
                        break
                rounds.append((oc, cc, date_label or f"Round {i+1}"))

            # Rows with actual negotiation data (below side_row + 1-2 header rows)
            data_start_ri = side_row_idx + 1
            # Skip extra sub-header rows (rep names, offer type labels)
            for ri in range(side_row_idx + 1, min(side_row_idx + 5, len(grid))):
                row = grid[ri]
                non_empty = [c for c in row if c]
                if non_empty and not any(_re.search(r'\d', c) for c in non_empty):
                    data_start_ri = ri + 1
                else:
                    break

            # ---------------------------------------------------------------- #
            # Smart label column detection
            # Score each pre-data column by how many rows have a descriptive
            # text label (not a pure number, not a date, contains letters).
            # ---------------------------------------------------------------- #
            def _is_descriptive_label(v: str) -> bool:
                if not v:
                    return False
                # Pure number → not a label
                if _re.match(r'^\d+(\.\d+)?$', v.strip()):
                    return False
                # ISO date → not a label
                if _re.search(r'\d{4}-\d{2}-\d{2}', v):
                    return False
                # Contains at least one letter → probably a label
                return bool(_re.search(r'[a-zA-Z]', v))

            label_col = 0
            best_label_score = -1
            max_label_ci = min(owner_cols[0] if owner_cols else 3, 5)
            for ci in range(0, max_label_ci):
                score = sum(
                    1 for ri in range(data_start_ri, len(grid))
                    if ci < len(grid[ri]) and _is_descriptive_label(grid[ri][ci])
                )
                if score > best_label_score:
                    best_label_score = score
                    label_col = ci

            # ---------------------------------------------------------------- #
            # Helper: is this row a real negotiation item (not metadata)?
            # Skip rows where the label looks like raw data, not a term name.
            # ---------------------------------------------------------------- #
            def _is_metadata_row(label: str, owner_val: str, bdd_val: str) -> bool:
                if not label:
                    return False
                l = label.strip().lower()
                # Pure number used as label
                if _re.match(r'^\d+(\.\d+)?$', label.strip()):
                    return True
                # ISO datetime as label
                if _re.search(r'\d{4}-\d{2}-\d{2}', label):
                    return True
                # Single generic word with no negotiation meaning
                if l in ('date', 'recommended', 'notes', 'remarks', 'comments', 'status', 'total',
                         'client', 'owner', 'tenant', 'lessor', 'lessee', 'location', 'venue',
                         'property', 'unit', 'floor', 'building', 'bldg', 'type', 'use'):
                    return True
                # Long person name (≥4 words, no digits) → owner name row
                words = label.split()
                if len(words) >= 4 and not _re.search(r'\d', label) and label == (owner_val or bdd_val or label):
                    return True
                # Looks like an address (contains "Street", "Ave", "Village", "Q.C", "City")
                if _re.search(r'\b(street|ave|avenue|village|q\.c|city|blvd|road|rd\.)\b', l):
                    return True
                # Meeting location/time row: contains time (am/pm) or meeting keywords, AND bdd is empty
                if not bdd_val and _re.search(r'\b(am|pm)\b|\b(office|meeting|conference)\b', l):
                    return True
                # Meeting type row: label is a meeting/activity type AND bdd is empty
                if not bdd_val and _re.search(r'\b(meeting|activity|call|visit|site|inspection)\b', l):
                    return True
                # Label is a company/brand/location name that is identical to (or contained in) the client value
                # and BDD has no position (e.g. "SOGO" title with BDD empty, client "SOGO")
                if not bdd_val and owner_val and l == owner_val.strip().lower():
                    return True
                # Label contains "/" and looks like a location+time ("Manhattan / office 2:00 pm")
                if '/' in label and _re.search(r'\b(am|pm|\d{1,2}:\d{2}|office|floor|bldg|building)\b', l):
                    return True
                return False

            lines.append("=== NEGOTIATION TABLE (left → right = oldest round → latest round) ===")
            lines.append("")

            round_headers = " | ".join(f"[{d}] Owner | BDD" for _, _, d in rounds)
            lines.append(f"{'ITEM':<35} | {round_headers} | LATEST OWNER | LATEST BDD")
            lines.append("-" * 120)

            for ri in range(data_start_ri, len(grid)):
                row = grid[ri]
                if not any(row):
                    continue

                owner_vals = [row[oc] for oc, _, _ in rounds]
                bdd_vals = [row[cc] for _, cc, _ in rounds]
                if not any(owner_vals) and not any(bdd_vals):
                    continue

                # Item label from best detected label column, then fallback to other pre-data cols
                item_label = row[label_col] if label_col < len(row) and row[label_col] else ""
                if not item_label:
                    for lc in range(0, max_label_ci):
                        if lc != label_col and lc < len(row) and _is_descriptive_label(row[lc]):
                            item_label = row[lc]
                            break
                # Last resort: use row number (AI instructed to skip these)
                if not item_label:
                    item_label = f"[Row {ri+1} — no label found]"

                latest_owner = next((row[oc] for oc, _, _ in reversed(rounds) if row[oc]), "")
                latest_bdd = next((row[cc] for _, cc, _ in reversed(rounds) if row[cc]), "")

                # Skip obvious metadata rows
                if _is_metadata_row(item_label, latest_owner, latest_bdd):
                    continue

                round_cells = " | ".join(
                    f"{row[oc] or '-':>12} | {row[cc] or '-':<12}"
                    for oc, cc, _ in rounds
                )
                lines.append(f"{item_label:<35} | {round_cells} | {latest_owner:<14} | {latest_bdd}")

            lines.append("")

            # ---------------------------------------------------------------- #
            # 4. LATEST POSITIONS summary — critical for agreed/negotiating
            # ---------------------------------------------------------------- #
            lines.append("=== LATEST POSITIONS SUMMARY (use this to determine agreed vs negotiating) ===")
            lines.append("Rule: if Latest Owner == Latest BDD → AGREED. If different → NEGOTIATING.")
            lines.append("")

            for ri in range(data_start_ri, len(grid)):
                row = grid[ri]
                if not any(row):
                    continue
                item_label = row[label_col] if label_col < len(row) and row[label_col] else ""
                latest_owner = next((row[oc] for oc, _, _ in reversed(rounds) if row[oc]), "")
                latest_bdd = next((row[cc] for _, cc, _ in reversed(rounds) if row[cc]), "")
                if not latest_owner and not latest_bdd:
                    continue
                if _is_metadata_row(item_label, latest_owner, latest_bdd):
                    continue
                match = "✓ AGREED" if latest_owner == latest_bdd and latest_owner else "✗ NEGOTIATING"
                lines.append(f"  {item_label:<35} | Owner: {latest_owner:<20} | BDD: {latest_bdd:<20} | {match}")

        else:
            # ---------------------------------------------------------------- #
            # Fallback: no paired structure — dump labelled rows
            # ---------------------------------------------------------------- #
            lines.append("=== SPREADSHEET DATA ===")
            header_row = next(
                (r for r in grid if sum(1 for c in r if c) >= 2), []
            )
            for ri, row in enumerate(grid):
                parts = []
                for ci, val in enumerate(row):
                    if val:
                        hdr = header_row[ci] if ci < len(header_row) and header_row[ci] else f"Col{ci+1}"
                        parts.append(f"{hdr}: {val}")
                if parts:
                    lines.append(" | ".join(parts))

        if lines:
            sheets_text.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(lines))

    wb.close()
    return "\n\n".join(sheets_text)


def _extract_xls_text(content: bytes) -> str:
    try:
        import xlrd
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Legacy .xls format requires xlrd. Please convert to .xlsx and re-upload.",
        )

    wb = xlrd.open_workbook(file_contents=content)
    sheets_text: list[str] = []

    for sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
        rows_text: list[str] = []

        for i in range(min(ws.nrows, 200)):
            cells = [
                str(ws.cell_value(i, j)) if j < ws.ncols else ""
                for j in range(min(ws.ncols, 27))
            ]
            if any(c.strip() for c in cells):
                rows_text.append("\t".join(cells))

        if rows_text:
            sheets_text.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows_text))

    return "\n\n".join(sheets_text)


def _extract_csv_text(content: bytes) -> str:
    try:
        decoded = content.decode("utf-8")
    except UnicodeDecodeError:
        decoded = content.decode("latin-1")

    reader = csv.reader(io.StringIO(decoded))
    rows: list[str] = []
    for i, row in enumerate(reader):
        if i > 200:
            break
        if any(cell.strip() for cell in row):
            rows.append("\t".join(row))

    return "\n".join(rows)


def _find_best_sheet(sheetnames: list, hint: str) -> str | None:
    """
    Given a property name hint (e.g. 'Maxx Hotel'), return the sheet name
    that best matches it, or None if no reasonable match is found.
    Matching is case-insensitive and checks for substring overlap.
    """
    if not hint:
        return None
    import re as _re
    # Normalise: lowercase, remove punctuation/extra spaces
    def _norm(s: str) -> str:
        return _re.sub(r'[^a-z0-9 ]', ' ', s.lower())

    hint_words = set(_norm(hint).split())
    best_sheet, best_score = None, 0

    for name in sheetnames:
        name_words = set(_norm(name).split())
        overlap = len(hint_words & name_words)
        if overlap > best_score:
            best_score = overlap
            best_sheet = name

    # Require at least one word in common (ignores generic words like "the")
    return best_sheet if best_score >= 1 else None


def _file_to_text(content: bytes, filename: str, property_name: str = "") -> str:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _extract_csv_text(content)
    elif lower.endswith(".xls"):
        return _extract_xls_text(content)
    else:
        return _extract_xlsx_text(content, property_name=property_name)


# --------------------------------------------------------------------------- #
# Gemini prompt                                                                #
# --------------------------------------------------------------------------- #

_SYSTEM_INSTRUCTION = (
    "You are a senior real estate negotiation analyst specialising in commercial property deals "
    "in the Philippines and Southeast Asia. "
    "Your job is to carefully read every cell of the spreadsheet and extract a structured, "
    "accurate negotiation DECISION TIMELINE — including whether each item is AGREED or still NEGOTIATING. "
    "CRITICAL: You MUST only extract information that is EXPLICITLY present in the spreadsheet data "
    "provided in the user message. Do NOT invent, assume, fabricate, or hallucinate ANY values, "
    "names, dates, figures, or negotiation terms. If data is not in the spreadsheet, do not include it. "
    "Always respond with valid JSON only — no markdown, no explanation, just the JSON object."
)

_USER_PROMPT_TEMPLATE = """\
Carefully analyse the following spreadsheet data and extract ALL negotiation information, \
focusing on building an accurate chronological DECISION TIMELINE.

⚠️  STRICT GROUNDING RULE: Every value, date, figure, and term you include in your output MUST \
come directly from the spreadsheet data below. Do NOT invent, assume, or extrapolate anything. \
If a field has no data in the spreadsheet, leave it null or omit it. \
If the spreadsheet content is insufficient to produce a meaningful analysis, set useful=false.

⚠️  DATE GROUNDING RULE (CRITICAL — violations will break the product):
- Dates you output MUST appear verbatim (or in an unambiguous equivalent form) in the spreadsheet text above.
- Do NOT guess, infer, or fill in dates from your training knowledge or pattern-matching.
- If the spreadsheet shows 2025 or 2026 dates, do NOT output 2023 or any other year not present.
- If you cannot find an explicit date in the spreadsheet, set the date field to null.
- Wrong dates are worse than null — prefer null over an invented date.

SPREADSHEET CONTENT:
{sheet_text}

=== HOW TO DETERMINE AGREEMENT vs NEGOTIATION ===

AGREED signals (mark event_type="agreement", outcome reflects the agreed value):
- Explicit words: "agreed", "confirmed", "accepted", "approved", "signed", "closed", "done", "OK", "✓", "checked"
- Phrases: "both parties agreed", "mutually accepted", "finalized", "no objection", "go signal", "proceed"
- A value that appears identical on both the BDD/broker side AND the client side with no further counter
- A date with a confirmation note beside it

STILL NEGOTIATING signals (mark event_type="offer" or "condition"):
- Words: "counter", "pending", "TBD", "for review", "under negotiation", "awaiting", "will revert", "to confirm"
- A value on one side with a different value on the other side (gap exists)
- Open questions, conditions not yet accepted, items marked "subject to"

OVERALL negotiationStatus rules:
- "Fully Agreed – Deal Closed" → every main term (price, payment, lease term) is confirmed by both parties
- "Partially Agreed – [specific items] still negotiating" → some items agreed, some not
- "Active – Counter offer stage" → price or key terms still have an open gap
- "Active – Awaiting client decision" → BDD has made a final offer, client has not responded
- "Stalled – No recent activity" → last entry is old with no resolution
- "Rejected – Deal fell through" → explicit rejection language found

=== OUTPUT SCHEMA ===

Return a JSON object with this exact schema:
{{
  "useful": boolean,
  "reason": "string (only if useful=false — explain why the file is not negotiation-related)",
  "negotiationStatus": "string — use one of the status formats above, be specific",
  "priceComparison": {{
    "askingPrice": "string (formatted, e.g. '₱150,000/month')",
    "clientOffer": "string (formatted)",
    "gap": "string (the difference, e.g. '₱20,000 or 13%') — use 'None – agreed' if price is settled",
    "currency": "string (e.g. 'PHP', 'USD')"
  }},
  "clientConditions": ["plain-English condition strings — include ALL conditions the client imposed"],
  "recentTransactions": [
    {{
      "date": "string or null (e.g. 'Feb 10, 2026')",
      "title": "string — short 3-6 word label (e.g. 'Cash Payment Agreed', 'Price Reduced to ₱12.5M', 'Client Rejected Offer')",
      "summary": "string — 1-3 sentences: WHO did WHAT, the exact values involved, and WHY",
      "event_type": "one of: payment_method | price_change | offer | agreement | condition | meeting | rejection | other",
      "outcome": "string or null — the confirmed result if agreed, or the next step if still open"
    }}
  ],
  "keyNotes": ["important flags, risks, or observations — include the overall deal status as the first note"],
  "rawExtracted": [
    {{ "header": "label/field name", "value": "the value" }}
  ]
}}

=== RULES ===
- Read EVERY row and column — do not skip data even if it appears peripheral.
- Set useful=false ONLY if the file has zero negotiation content (e.g. pure inventory list, user data).
- recentTransactions MUST be sorted chronologically (oldest first).
- Every recentTransaction MUST have a non-empty title and summary.
- rawExtracted must contain ALL negotiation-relevant key-value pairs found in the sheet.
- All monetary values must preserve their original currency, units, and formatting.
- Decimal values that represent rates or percentages (e.g. 0.04, 0.05) must be displayed as percentages (e.g. 4%, 5%).
- Omit fields that have absolutely no data (do not include null or empty fields).
- clientConditions and keyNotes must be plain-English strings, one point per item.
- Be specific — do not write vague summaries like "price was discussed". State the actual figures.
"""


_COMPARISON_SYSTEM_INSTRUCTION = (
    "You are a senior real estate negotiation analyst specialising in commercial property deals "
    "in the Philippines and Southeast Asia. "
    "Your job is to carefully read every cell of the spreadsheet and produce an accurate "
    "per-item comparison showing the BDD Employee position vs the Client position, "
    "and whether each item has been AGREED or is still NEGOTIATING. "
    "CRITICAL: You MUST only extract information that is EXPLICITLY present in the spreadsheet data "
    "provided in the user message. Do NOT invent, assume, fabricate, or hallucinate ANY values, "
    "names, dates, figures, or negotiation terms. If data is not in the spreadsheet, do not include it. "
    "Always respond with valid JSON only — no markdown, no explanation, just the JSON object."
)

_COMPARISON_PROMPT_TEMPLATE = """\
Carefully analyse the following real estate negotiation spreadsheet. \
Extract a precise per-item comparison between the broker/BDD Employee and the client (buyer/lessee).

⚠️  STRICT GROUNDING RULE: Every value, date, figure, and term you include in your output MUST \
come directly from the spreadsheet data below. Do NOT invent, assume, or extrapolate anything. \
If a field has no data in the spreadsheet, leave it null or omit it. \
If the spreadsheet content is insufficient to produce a meaningful analysis, set useful=false.

⚠️  DATE GROUNDING RULE (CRITICAL — violations will break the product):
- Dates you output MUST appear verbatim (or in an unambiguous equivalent form) in the spreadsheet text above.
- Do NOT guess, infer, or fill in dates from your training knowledge or pattern-matching.
- If the spreadsheet shows 2025 or 2026 dates, do NOT output 2023 or any other year not present.
- If you cannot find an explicit date in the spreadsheet for an item, set agreed_date / date to null.
- Wrong dates are worse than null — prefer null over an invented date.

IMPORTANT — HOW TO READ THE DATA:
- Columns are ordered LEFT → RIGHT = OLDEST round → LATEST/MOST RECENT round.
- Each negotiation round has two columns: [Owner/Client Side] and [Company/BDD Side].
- Always base bdd_offer and client_offer on the LATEST (rightmost) non-empty values for each item.
- The "LATEST POSITIONS SUMMARY" section (if present) already shows you the final positions — use it.
- A "✓ AGREED" marker in the summary means the values match — set status="agreed".
- A "✗ NEGOTIATING" marker means the values differ — set status="negotiating".

SPREADSHEET CONTENT:
{sheet_text}

=== HOW TO DETERMINE AGREED vs NEGOTIATING ===

Set status="agreed" when ANY of the following is true:
  1. **MATCHING VALUES (highest priority)** — the BDD offer value and the client offer value are \
the same amount/term (e.g. both say "₱500,000", both say "3 years", both say "Cash"). \
Matching values = deal reached, regardless of whether a confirmation word is present.
  2. One side explicitly accepted the other's exact value with no further counter \
(e.g. client says "OK with ₱500,000" when BDD offered ₱500,000).
  3. Explicit confirmation words are present for that item: "agreed", "confirmed", "accepted", \
"OK", "✓", "signed", "go signal", "finalized", "no objection", "proceed", "both parties confirmed".

Set status="negotiating" when ANY of the following is true:
  1. The BDD offer value and the client offer value are DIFFERENT (a gap exists).
  2. Only one side has stated a position with no response from the other.
  3. Words like "counter", "pending", "TBD", "awaiting", "for review", "subject to", \
"will revert", "under consideration" are present.
  4. The item is marked with a question mark or left blank on one side.

MATCHING VALUE EXAMPLES:
  - BDD: "₱500,000/month"  |  Client: "₱500,000/month"  →  status="agreed", final_value="₱500,000/month"
  - BDD: "3 years"          |  Client: "3 years"          →  status="agreed", final_value="3 years"
  - BDD: "₱800,000"         |  Client: "₱750,000"         →  status="negotiating" (different values)
  - BDD: "Cash"             |  Client: "Cash"             →  status="agreed", final_value="Cash"

=== OUTPUT SCHEMA ===

Return a JSON object with this exact schema:
{{
  "useful": boolean,
  "reason": "string (only if useful=false — explain why the file is not negotiation-related)",
  "overall_status": "string — accurately summarise: how many items are agreed vs negotiating, \
e.g. 'Deal Closed – All terms agreed', '4 of 7 items agreed, costing and payment schedule still open', \
'All terms still under negotiation'",
  "summary": "string — 2-4 sentence overview of where the deal stands, what has been settled, \
and what is still blocking progress",
  "negotiation_items": [
    {{
      "title": "string — the negotiation topic (e.g. 'Costing', 'Payment Method', 'Lease Term', \
'Security Deposit', 'Advance Rent', 'Escalation Rate', 'Move-in Date')",
      "bdd_offer": "string — the exact BDD/broker position or offer with values",
      "client_offer": "string — the exact client counter-offer or position with values",
      "status": "agreed | negotiating",
      "agreed_date": "string or null — The date when both sides converged on the same value. \
Look for the date column/row in the spreadsheet where both BDD and client show matching values. \
If the client proposed a value on 05/27 and BDD matched it on 05/28, agreed_date = '05/28'. \
Use the LATER of the two dates (i.e. when the second party accepted). \
IMPORTANT: set to null unless a specific date is EXPLICITLY VISIBLE in the spreadsheet text — \
do NOT infer, guess, or fabricate a date. A null is always better than an invented date.",
      "final_value": "string or null — the exact agreed value (ONLY when status=agreed)",
      "notes": "string or null — important context, conditions, or blockers for this item"
    }}
  ]
}}

=== RULES ===
- Read EVERY row and column — do not skip any negotiation topic.
- Extract ALL negotiation topics — for LEASE deals this includes but is not limited to: \
Monthly Rent, Area Rate, Security Deposit, Advance Rent, Lease Term, Escalation Rate, Escalation Start, \
Payment Terms, Grace Period, Fit-out Period, Free Rental, Parking, Renovation Budget, Move-in Date, \
Orientation, Commission. \
For SALE deals this includes but is not limited to: Sale Price, Total Contract Price, Payment Scheme, \
Down Payment, Spot Down Payment, Installment Terms, Balance Payment, Tax Allocation, Transfer of Title, \
Commission, Move-in Date. \
Include ANY row where the Owner/Client column OR the BDD/Company column contains a specific value, \
figure, rate, term, or condition — even if the label does not appear in the examples above.
- **ONE ROW PER TOPIC** — Each negotiation topic must appear EXACTLY ONCE in negotiation_items. \
Do NOT create separate rows for "Initial Offer" and "Latest Position" of the same topic. \
Do NOT append "(Initial Offer)", "(Latest Position)", "(Round 1)", etc. to the title. \
The title must be the plain topic name only (e.g. "Security Deposit", not "Security Deposit (Latest Position)"). \
If a topic appears multiple times across rounds, consolidate into ONE row using the LATEST values only.
- **TITLE MUST BE THE NEGOTIATION TERM NAME** — Use the descriptive label from the label/term column \
(e.g. "Area Rate", "Security Deposit", "Lease Term", "Advance Rent", "Escalation Rate"). \
NEVER use "Row X", raw numbers, spreadsheet row identifiers, or "[Row X — no label found]" as a title — \
if a row has no meaningful label, SKIP it entirely.
- **SKIP METADATA ROWS** — Do NOT include rows for:
  * Owner/person names or property addresses
  * Raw date values as topics ("2021-07-13", "Date")
  * Single generic words ("Recommended", "Status")
  * Meeting logistics: meeting location ("Manhattan / office 2:00 pm"), meeting type \
("Personal Meeting", "Site Visit"), venue/hotel names used as row labels ("SOGO", "Ortigas"), \
or activity/note rows — these are context info, NOT negotiation terms
  * Any row where the BDD Employee column is empty AND the value is just a location, \
meeting type, or company name
  * Row labels that are property/deal identifiers: "Client", "Owner", "Tenant", \
"Location", "Property", "Venue", "Unit", "Floor", "Building", "Type"
  * Row labels that are note fields: "Notes", "Remarks", "Comments"
  * INCLUDE every row where either the Owner/Client column or the BDD/Company column \
contains a value (a number, amount, rate, term length, condition, or option). \
Even if a row label is unfamiliar, if it has negotiated values on both sides (or at least one side), \
include it. \
  * EXCLUDE a row ONLY when it is pure metadata with NO negotiated values — \
e.g. rows where both sides are empty, or where the row is just a section divider or category header \
with no values at all.
- status must be exactly "agreed" or "negotiating" — no other values allowed.
- agreed_date: for agreed items, find the date in the spreadsheet when both values matched. \
  Use the later date (when the second party accepted). If the spreadsheet has date columns or \
  date rows, extract the exact date. Set null only if truly no date is present anywhere.
- final_value is ONLY non-null when status="agreed".
- negotiation_items must be a list — never null or omitted.
- All monetary values must preserve their original currency, units, and formatting.
- Decimal values that represent rates or percentages (e.g. 0.04, 0.05) must be displayed as percentages (e.g. 4%, 5%).
- Be specific — include exact figures (e.g. "₱850,000/month" not just "high price").
- overall_status must honestly reflect the data — do not over-optimistically say "agreed" \
if any key term is unresolved.
"""


async def _call_gemini_comparison(sheet_text: str, api_key: str, structure_guide: str = "") -> AIComparisonResult:
    """Call Gemini REST API with the comparison prompt and return parsed AIComparisonResult."""
    import httpx

    full_text = (structure_guide + sheet_text) if structure_guide else sheet_text
    print(f"📄 Sending {len(full_text)} chars to Gemini comparison. Preview:\n{full_text[:800]}\n---")

    gemini_url = (
        "https://generativelanguage.googleapis.com/v1beta/models"
        f"/gemini-2.5-pro:generateContent?key={api_key}"
    )
    payload = {
        "system_instruction": {"parts": [{"text": _COMPARISON_SYSTEM_INSTRUCTION}]},
        "contents": [
            {"parts": [{"text": _COMPARISON_PROMPT_TEMPLATE.format(sheet_text=full_text)}]}
        ],
        "generationConfig": {
            "temperature": 0.1,
            "responseMimeType": "application/json",
            "maxOutputTokens": 65536,
            # thinkingBudget: Gemini 2.5 Pro extended thinking helps it reason through
            # complex multi-round negotiation tables.  24 000 tokens is enough for deep
            # reasoning on most sheets without hitting the 32 000 cap.
            "thinkingConfig": {"thinkingBudget": 24000},
        },
    }

    async with httpx.AsyncClient(timeout=300.0) as http_client:
        gemini_response = await http_client.post(gemini_url, json=payload)

    if gemini_response.status_code != 200:
        print(f"💥 Gemini API error {gemini_response.status_code}: {gemini_response.text[:500]}")
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Gemini API error {gemini_response.status_code}: {gemini_response.text[:300]}",
        )

    gemini_data = gemini_response.json()
    candidate = gemini_data["candidates"][0]
    finish_reason = candidate.get("finishReason", "STOP")
    if finish_reason == "MAX_TOKENS":
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="AI response was truncated (output token limit hit). Please try with a smaller file.",
        )
    result_text = candidate["content"]["parts"][0]["text"]
    result_dict = json.loads(result_text)
    # Gemini occasionally wraps the object in an array — unwrap it
    if isinstance(result_dict, list):
        result_dict = result_dict[0] if result_dict else {}
    return AIComparisonResult(**result_dict)


async def _call_gemini(sheet_text: str, api_key: str) -> AIAnalysisResult:
    """Call Gemini REST API and return parsed AIAnalysisResult."""
    import httpx

    print(f"📄 Sending {len(sheet_text)} chars to Gemini analysis. Preview:\n{sheet_text[:800]}\n---")

    gemini_url = (
        "https://generativelanguage.googleapis.com/v1beta/models"
        f"/gemini-2.5-pro:generateContent?key={api_key}"
    )
    payload = {
        "system_instruction": {"parts": [{"text": _SYSTEM_INSTRUCTION}]},
        "contents": [
            {"parts": [{"text": _USER_PROMPT_TEMPLATE.format(sheet_text=sheet_text)}]}
        ],
        "generationConfig": {
            "temperature": 0.1,
            "responseMimeType": "application/json",
            "maxOutputTokens": 65536,
            "thinkingConfig": {"thinkingBudget": 24000},
        },
    }

    async with httpx.AsyncClient(timeout=300.0) as http_client:
        gemini_response = await http_client.post(gemini_url, json=payload)

    if gemini_response.status_code != 200:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Gemini API error {gemini_response.status_code}: {gemini_response.text[:300]}",
        )

    gemini_data = gemini_response.json()
    candidate = gemini_data["candidates"][0]
    finish_reason = candidate.get("finishReason", "STOP")
    if finish_reason == "MAX_TOKENS":
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="AI response was truncated (output token limit hit). Please try with a smaller file.",
        )
    result_text = candidate["content"]["parts"][0]["text"]
    result_dict = json.loads(result_text)
    # Gemini occasionally wraps the object in an array — unwrap it
    if isinstance(result_dict, list):
        result_dict = result_dict[0] if result_dict else {}
    return AIAnalysisResult(**result_dict)


# --------------------------------------------------------------------------- #
# Endpoints                                                                    #
# --------------------------------------------------------------------------- #

@router.post("/analyze", response_model=AIAnalysisResult)
async def analyze_negotiation_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """
    Analyse an uploaded negotiation file (xlsx, xls, csv) using Google Gemini AI.
    Returns structured negotiation insights — does NOT persist to database.
    """
    api_key = settings.GOOGLE_GEMINI_API_KEY
    if not api_key:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI analysis is not configured on this server (missing GOOGLE_GEMINI_API_KEY).",
        )

    filename = file.filename or ""
    lower = filename.lower()
    if not any(lower.endswith(ext) for ext in (".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload .xlsx, .xls, or .csv",
        )

    content = await file.read()

    try:
        sheet_text = _file_to_text(content, filename)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not read file: {exc}",
        )

    if not sheet_text.strip():
        return AIAnalysisResult(useful=False, reason="The uploaded file appears to be empty.")

    try:
        ai_result = await _call_gemini(sheet_text, api_key)
        ai_result = _strip_hallucinated_transaction_dates(ai_result, sheet_text)
        return ai_result
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to parse AI response: {exc}",
        )
    except Exception as exc:
        print(f"💥 Gemini analysis error: {exc}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI analysis failed: {exc}",
        )


@router.post("/upload/{nego_table_id}", response_model=UploadAnalysisResponse)
async def upload_and_analyze(
    nego_table_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(get_current_user),
):
    """
    Upload a negotiation file, run Gemini AI analysis, auto-save to DB, and return the result.
    This is the single-step endpoint used by the new Negotiation Timeline UI.
    """
    api_key = settings.GOOGLE_GEMINI_API_KEY
    if not api_key:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI analysis is not configured on this server (missing GOOGLE_GEMINI_API_KEY).",
        )

    # Validate nego table exists in DB
    result = await db.execute(select(NegoTable).filter(NegoTable.id == nego_table_id))
    nego_table = result.scalar_one_or_none()
    if not nego_table:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Negotiation table {nego_table_id} not found",
        )

    filename = file.filename or ""
    lower = filename.lower()
    if not any(lower.endswith(ext) for ext in (".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload .xlsx, .xls, or .csv",
        )

    content = await file.read()

    # Parse spreadsheet to text
    try:
        sheet_text = _file_to_text(content, filename)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not read file: {exc}",
        )

    if not sheet_text.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded file appears to be empty.",
        )

    # Call Gemini AI
    try:
        ai_result = await _call_gemini(sheet_text, api_key)
        ai_result = _strip_hallucinated_transaction_dates(ai_result, sheet_text)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to parse AI response: {exc}",
        )
    except Exception as exc:
        print(f"💥 Gemini upload+analyze error: {exc}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI analysis failed: {exc}",
        )

    # Upload file to OSS storage
    await file.seek(0)
    try:
        upload_result = await file_storage_service.save_file(file, subfolder="negotiation_chronicles")
        file_url = upload_result["secure_url"]
    except Exception as exc:
        print(f"⚠️  OSS upload failed, storing without file URL: {exc}")
        file_url = ""

    # Build rawExtracted for backward-compat parsed_data column
    raw_extracted = ai_result.rawExtracted or []
    parsed_data = [{"header": r.header, "value": r.value} for r in raw_extracted]

    # Persist to database
    attachment = NegotiationChronicleAttachment(
        nego_table_id=nego_table_id,
        filename=filename,
        file_url=file_url,
        file_type=lower.rsplit(".", 1)[-1],
        file_size=len(content),
        parsed_data=parsed_data,
        ai_result=ai_result.model_dump(),
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    await db.commit()
    await db.refresh(attachment)

    print(f"✅ Saved AI timeline attachment {attachment.id} for nego table {nego_table_id}")

    return UploadAnalysisResponse(
        attachment_id=attachment.id,
        ai_result=ai_result,
    )


@router.post("/upload-for-property/{property_id}", response_model=UploadForPropertyResponse)
async def upload_for_property(
    property_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(get_current_user),
):
    """
    Upload a negotiation file for a property. Auto-creates a NegoTable if one doesn't exist.
    Runs Gemini AI comparison analysis (BDD Employee vs Client per item) and saves to DB.
    Only BDD_USER (assigned reviewer) and ADMIN may upload.
    """
    from app.models.nego_table import NegoTableStatus

    api_key = settings.GOOGLE_GEMINI_API_KEY
    if not api_key:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI analysis is not configured on this server (missing GOOGLE_GEMINI_API_KEY).",
        )

    filename = file.filename or ""
    lower = filename.lower()
    if not any(lower.endswith(ext) for ext in (".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload .xlsx, .xls, or .csv",
        )

    content = await file.read()

    # Load property name upfront so we can target the correct sheet in multi-sheet workbooks
    prop_result = await db.execute(select(Property).where(Property.id == property_id))
    prop = prop_result.scalar_one_or_none()
    prop_name = prop.name if prop else ""

    # Parse spreadsheet to text — pass property name so only the matching sheet is extracted
    try:
        sheet_text = _file_to_text(content, filename, property_name=prop_name)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not read file: {exc}",
        )

    if not sheet_text.strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded file appears to be empty.",
        )

    # Build structure guide so the AI understands column layout before reading the raw dump
    structure_guide = _detect_sheet_structure(content, filename, target_sheet_hint=prop_name)

    # Call Gemini AI with comparison prompt first — before touching the DB
    try:
        ai_result = await _call_gemini_comparison(sheet_text, api_key, structure_guide)
        ai_result = _deduplicate_items(ai_result)
        ai_result = _enforce_agreed_on_match(ai_result)
        ai_result = _strip_hallucinated_dates(ai_result, sheet_text)
    except HTTPException:
        raise  # let 502/503 from Gemini propagate with its real detail
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to parse AI response: {exc}",
        )
    except Exception as exc:
        print(f"💥 Gemini comparison error: {type(exc).__name__}: {exc}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI analysis failed: {type(exc).__name__}: {exc}",
        )

    # If Gemini says the file is not negotiation-related, return early — nothing is saved
    if not ai_result.useful:
        print(f"⚠️  File '{filename}' rejected by AI for property {property_id}: {ai_result.reason}")
        return UploadForPropertyResponse(
            attachment_id=-1,
            ai_result=ai_result,
            nego_table_id=-1,
        )

    # Get or create NegoTable for this property (only reached when file is useful)
    result = await db.execute(select(NegoTable).where(NegoTable.property_id == property_id))
    nego_table = result.scalar_one_or_none()
    if not nego_table:
        # prop was already loaded above for sheet matching — reuse it
        prop_name = prop.name if prop else "Unknown"
        prop_address = prop.address if prop else "Unknown"
        prop_type = str(prop.property_type.value) if prop else "Unknown"
        prop_lot_area = float(prop.lot_area) if prop else 0.0

        nego_table = NegoTable(
            property_id=property_id,
            status=NegoTableStatus.ACTIVE,
            created_by_id=current_user.id,
            referred_date=datetime.utcnow(),
            source_origin="Upload",
            original_property_name=prop_name,
            current_property_name=prop_name,
            original_location=prop_address,
            current_location=prop_address,
            original_property_type=prop_type,
            current_property_type=prop_type,
            original_lot_area=prop_lot_area,
            current_lot_area=prop_lot_area,
        )
        db.add(nego_table)
        await db.commit()
        await db.refresh(nego_table)
        print(f"✅ Auto-created NegoTable {nego_table.id} for property {property_id}")

    # Replace existing attachments — keep only the latest per NegoTable
    old_attachments_result = await db.execute(
        select(NegotiationChronicleAttachment).where(
            NegotiationChronicleAttachment.nego_table_id == nego_table.id
        )
    )
    for old in old_attachments_result.scalars().all():
        await db.delete(old)
    await db.commit()
    print(f"🗑️  Cleared old attachments for nego table {nego_table.id}")

    # Upload file to OSS storage
    await file.seek(0)
    try:
        upload_result = await file_storage_service.save_file(file, subfolder="negotiation_chronicles")
        file_url = upload_result["secure_url"]
    except Exception as exc:
        print(f"⚠️  OSS upload failed, storing without file URL: {exc}")
        file_url = ""

    # Persist attachment with ai_result
    attachment = NegotiationChronicleAttachment(
        nego_table_id=nego_table.id,
        filename=filename,
        file_url=file_url,
        file_type=lower.rsplit(".", 1)[-1],
        file_size=len(content),
        parsed_data=[],
        ai_result=ai_result.model_dump(),
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    await db.commit()
    await db.refresh(attachment)

    print(f"✅ Saved AI comparison attachment {attachment.id} for property {property_id} (nego table {nego_table.id})")

    return UploadForPropertyResponse(
        attachment_id=attachment.id,
        ai_result=ai_result,
        nego_table_id=nego_table.id,
    )


@router.get("/attachments-for-property/{property_id}", response_model=List[NegotiationChronicleAttachmentSchema])
async def get_attachments_for_property(
    property_id: int,
    db: AsyncSession = Depends(get_async_session),
    current_user: User = Depends(get_current_user),
):
    """
    Get all negotiation attachments for a property directly by property_id.
    Bypasses the in-memory nego_tables_simple router — queries the real DB.
    Used by the property detail page on every load/refresh.
    """
    # Find NegoTable for this property in the DB
    result = await db.execute(select(NegoTable).where(NegoTable.property_id == property_id))
    nego_table = result.scalar_one_or_none()
    if not nego_table:
        return []

    # Return all attachments ordered newest first
    attachments_result = await db.execute(
        select(NegotiationChronicleAttachment)
        .where(NegotiationChronicleAttachment.nego_table_id == nego_table.id)
        .order_by(NegotiationChronicleAttachment.created_at.desc())
    )
    return attachments_result.scalars().all()
