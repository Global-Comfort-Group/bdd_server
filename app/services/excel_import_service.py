"""
Excel import service for BDD Property Tracker.

Parses the BDD monthly "REFERRED PROPERTIES Coding" Excel file and extracts
property rows from all monthly sheets for bulk import.

Column positions are resolved PER SHEET from the header text, not hardcoded.
The workbook uses (at least) two different layouts:

  Jan, Aug-Dec:  C=PROPERTY NAME  D=PROPERTY CODING  E=REFERRED BY  F=Name
                 G=Property Sourcing ... M=STATUS

  Feb-Jul:       C=PROPERTY NAME  D=PROPERTY CODING  E=Lot Area  F=CFA
                 G=Lease  H=Sale  I=REFERRED BY  J=Name  K=Property Sourcing

An earlier version of this parser hardcoded the first layout's letters, so on
Feb-Jul sheets it read `CFA (Sqm.)` as the referrer name and a tally column as
the status. Resolving by header text fixes both and picks up the lot area,
floor area and lease/sale columns that the fixed mapping ignored.
"""
from io import BytesIO
from typing import Any, Dict, List, Optional
import re

import openpyxl

# Sheets to skip — they are aggregate/summary sheets, not property data
_SKIP_SHEETS = {"Jan-Nov. Summary", "Summary"}

# Header cell value that marks the column header row in each monthly sheet
_HEADER_MARKER = "PROPERTY NAME"

# Header text -> logical field. Matching is done on a normalized form
# (uppercase, alphanumerics only) against the START of the header, so
# "Lot Area (Sqm)" and "Lot Area (Sqm.)" both hit "LOTAREA".
#
# Note "REFERREBY BY" — the header is misspelled in the source file, so it is
# matched by a prefix short enough to survive the typo.
# ORDER MATTERS — the first prefix that matches wins, so a longer, more
# specific prefix must come before a shorter one it starts with
# ("LEASE PRICE" before "Lease", or the template's price column would be read
# as the monthly sheets' lease rate).
_HEADER_FIELDS = [
    ("PROPERTYNAME", "name"),
    ("PROPERTYCODING", "address"),
    ("PROPERTYSOURCING", "sourced_by"),
    ("PROPERTYTYPE", "property_type"),
    ("LOTAREA", "lot_area"),
    ("CFA", "building_area"),
    ("LEASEPRICE", "lease_price"),  # template — a number
    ("LEASE", "lease_raw"),         # monthly sheets — free text ("300/sqm.")
    ("SALE", "sale_raw"),
    ("PRICE", "price"),
    ("ZONING", "zoning_classification"),
    ("TRANSACTION", "transaction_status"),
    ("TITLENUMBER", "title_number"),
    ("FLOORS", "floors"),
    ("ROOMS", "rooms"),
    ("PARKING", "parking_slots"),
    ("DESCRIPTION", "description"),
    ("REFERRE", "referral_type"),   # "REFERRED BY" / "REFERREBY BY"
    ("STATUS", "status_hint"),
]

# Fields the complete-format template supplies that the monthly sheets do not.
# Absent on a monthly sheet, which simply leaves them None.
_TEMPLATE_TEXT_FIELDS = (
    "property_type", "zoning_classification", "transaction_status", "title_number",
    "description",
)
_TEMPLATE_NUMBER_FIELDS = ("price", "lease_price", "floors", "rooms", "parking_slots")


def _norm_header(value: Any) -> str:
    """Uppercase, alphanumerics only — so punctuation and spacing in the
    header text cannot affect matching."""
    if value is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def _cell_str(cell) -> Optional[str]:
    """Return stripped string value of a cell, or None if empty."""
    if cell is None or cell.value is None:
        return None
    # Non-breaking spaces appear in this workbook's address column.
    val = str(cell.value).replace("\xa0", " ").strip()
    return val if val else None


def _parse_number(value: Optional[str]) -> Optional[float]:
    """Parse an area cell to a float.

    A trailing unit is stripped when it restates the column's own unit
    ("800 sqm." -> 800.0). Anything else returns None rather than a guess:
    the column really does contain multi-lot values ("1,250 & 1,500",
    "1383.1 / 1,300", "Lot 1 - 1,745 sqm Lot 2 - 1,570 sqm") and the
    occasional different unit ("8.2 hec."). Collapsing those to one number
    would silently misstate a property's area, so they are left for a human.

    Note the caller treats None as "no data" and stores 0.0, so an
    unparseable area is indistinguishable from a missing one.
    """
    if not value:
        return None
    text = value.replace(",", "").strip()
    # Trailing "sqm" / "sq.m." / "sqm." — same unit as the column header.
    text = re.sub(r"\s*sq\.?\s*m\.?\s*$", "", text, flags=re.IGNORECASE).strip()
    if not re.fullmatch(r"\d+(\.\d+)?", text):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _find_header_row(sheet) -> Optional[int]:
    """Scan for the row where column C == 'PROPERTY NAME'.
    Returns the 1-based row index, or None if not found."""
    for row in sheet.iter_rows():
        for cell in row:
            # read_only mode may return EmptyCell objects — guard with hasattr
            if not hasattr(cell, "column") or cell.column != 3:
                continue
            if _norm_header(cell.value) == _norm_header(_HEADER_MARKER):
                return cell.row
    return None


def _build_column_map(sheet, header_row: int) -> Dict[str, int]:
    """Map logical field name -> 1-based column index, for THIS sheet.

    The referrer's name lives in a 'Name' column immediately after
    'REFERRED BY'. 'Name' is too generic to match on its own — the
    'Property Sourcing' column holds names too, of internal staff rather than
    the referrer — so it is anchored positionally off the REFERRED BY column.
    """
    columns: Dict[str, int] = {}
    headers: Dict[int, str] = {}

    for cell in sheet[header_row]:
        if not hasattr(cell, "column"):
            continue
        norm = _norm_header(cell.value)
        if not norm:
            continue
        headers[cell.column] = norm
        for prefix, field in _HEADER_FIELDS:
            if field in columns:
                continue
            if norm.startswith(prefix):
                columns[field] = cell.column
                break

    # referred_by = the "Name" column directly right of "REFERRED BY"
    ref_col = columns.get("referral_type")
    if ref_col is not None and headers.get(ref_col + 1, "").startswith("NAME"):
        columns["referred_by"] = ref_col + 1

    return columns


def parse_excel_properties(file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parse a BDD monthly Excel file and extract all property rows.

    Returns a list of dicts, each with:
        row_id         — unique string key: "<sheet_name>|<row_number>"
        sheet_name     — name of the monthly sheet
        row_number     — 1-based row index in the sheet
        name           — property name
        address        — property coding / address
        referred_by    — referrer name
        referral_type  — Broker / Council / BDD / Employee, when present
        lot_area       — float, when the sheet has the column and it parses
        building_area  — float (CFA), same conditions
        lease_raw      — raw lease text, e.g. "300/sqm." (see note below)
        sale_raw       — raw sale text, e.g. "94Million"
        status_hint    — raw status text, may be None

    lease_raw/sale_raw are deliberately NOT parsed into a price. That column
    mixes per-square-metre rates ("35,000/sqm.") with absolute totals
    ("94Million"); collapsing both into one numeric field would silently
    invent wrong prices. They are carried through as text, and their PRESENCE
    determines transaction_status at import time.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    results: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue

        sheet = wb[sheet_name]
        header_row = _find_header_row(sheet)
        if header_row is None:
            continue  # Sheet has no recognisable header — skip

        columns = _build_column_map(sheet, header_row)
        name_col = columns.get("name")
        address_col = columns.get("address")
        if name_col is None or address_col is None:
            continue  # Without these two there is nothing importable

        def col(cells, field):
            idx = columns.get(field)
            return _cell_str(cells.get(idx)) if idx else None

        for row in sheet.iter_rows(min_row=header_row + 1):
            # Guard against EmptyCell objects returned in read_only mode
            cells = {cell.column: cell for cell in row if hasattr(cell, "column")}

            name = _cell_str(cells.get(name_col))
            address = _cell_str(cells.get(address_col))

            # Skip rows with no name AND no address (blank / separator rows)
            if not name and not address:
                continue

            # Also skip rows where name looks like a number only (row counters)
            if name and name.isdigit():
                continue

            # row[0] can be an EmptyCell in read_only mode, which has no .row —
            # that happens whenever column A of the row is blank. Take the
            # number from any real cell instead.
            row_number = next(
                (c.row for c in row if hasattr(c, "row") and c.row is not None), 0
            )

            # The template sheet carries the fields the monthly sheets lack. On
            # a monthly sheet these columns simply do not exist, so every one
            # resolves to None and the lead still needs a human at promotion.
            extra: Dict[str, Any] = {
                f: col(cells, f) for f in _TEMPLATE_TEXT_FIELDS
            }
            for f in _TEMPLATE_NUMBER_FIELDS:
                extra[f] = _parse_number(col(cells, f))
            # Enum-ish values are normalized here so "hotel" or "Land & Building"
            # from a hand-typed cell still match; anything unrecognised is left
            # as-is for the schema to reject rather than silently corrected.
            if extra.get("property_type"):
                extra["property_type"] = (
                    extra["property_type"].strip().upper().replace(" & ", "_AND_").replace(" ", "_")
                )
            if extra.get("transaction_status"):
                extra["transaction_status"] = extra["transaction_status"].strip().upper()

            results.append({
                "row_id": f"{sheet_name}|{row_number}",
                "sheet_name": sheet_name,
                "row_number": row_number,
                "name": name,
                "address": address,
                "referred_by": col(cells, "referred_by"),
                "referral_type": col(cells, "referral_type"),
                "lot_area": _parse_number(col(cells, "lot_area")),
                "building_area": _parse_number(col(cells, "building_area")),
                "lease_raw": col(cells, "lease_raw"),
                "sale_raw": col(cells, "sale_raw"),
                "status_hint": col(cells, "status_hint"),
                **extra,
            })

    wb.close()
    return results
